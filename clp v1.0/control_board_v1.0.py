import sys
import os
import datetime
import sqlite3
import random
import string
import win32com.client as win32
import time
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QLineEdit,
    QVBoxLayout,
    QMessageBox,
    QDialog,
    QHBoxLayout,
    QFileDialog,
    QListWidget,
    QMainWindow,
    QListWidgetItem,
    QComboBox,
    QHeaderView,
    QGroupBox,
    QCheckBox,
    QFormLayout,
    QDateEdit,
    QSpinBox,
    QGraphicsOpacityEffect,
)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QFont, QPixmap, QIcon
from collections import defaultdict
from tkinter import simpledialog, messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Função para retornar o caminho completo do ícone
def icon_path(icon_name):
    dir_path = os.path.dirname(os.path.realpath(__file__))
    return os.path.join(dir_path, icon_name)


class DatabaseManager:
    def __init__(self):
        self.conn = sqlite3.connect("BDCALL.db")
        self.cursor = self.conn.cursor()
        self.create_table()

    def create_table(self):
        self.cursor.execute(
            """
        CREATE TABLE IF NOT EXISTS lavagem_placas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT,
            turno TEXT,
            hora TEXT,
            modelo TEXT,
            responsavel TEXT,
            linha_solicitante TEXT,
            blank_id TEXT,
            serial TEXT,
            fase TEXT
        )
        """
        )

        self.cursor.execute(
            """
    CREATE TABLE IF NOT EXISTS log_board (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        table_name TEXT,
        record_id INTEGER,
        field_name TEXT,
        old_value TEXT,
        new_value TEXT,
        date TEXT,
        time TEXT,
        serial TEXT
    )
    """
        )

        # Nova tabela ID_MODELO
        self.cursor.execute(
            """
        CREATE TABLE IF NOT EXISTS ID_MODELO (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            modelo TEXT
        )
        """
        )

        # Nova tabela ID_USER
        self.cursor.execute(
            """
        CREATE TABLE IF NOT EXISTS ID_USER (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT
        )
        """
        )
        self.conn.commit()

    # Métodos para interagir com a nova tabela ID_MODELO
    def insert_modelo(self, modelo):
        self.cursor.execute(
            """
        INSERT INTO ID_MODELO (modelo)
        VALUES (?)
        """,
            (modelo,),
        )
        self.conn.commit()

    def fetch_all_modelos(self):
        self.cursor.execute(
            """
        SELECT * FROM ID_MODELO
        """
        )
        return self.cursor.fetchall()

    def delete_modelo(self, modelo_id):
        self.cursor.execute(
            """
        DELETE FROM ID_MODELO WHERE id = ?
        """,
            (modelo_id,),
        )
        self.conn.commit()

    def fetch_all_responsaveis(self):
        self.cursor.execute(
            """
        SELECT * FROM ID_USER
        """
        )
        return self.cursor.fetchall()

    def update_user(self, old_nome, new_nome):
        self.cursor.execute(
            "UPDATE ID_USER SET nome = ? WHERE nome = ?", (new_nome, old_nome)
        )
        self.conn.commit()

    def delete_user(self, nome):
        self.cursor.execute("DELETE FROM ID_USER WHERE nome = ?", (nome,))
        self.conn.commit()

    def insert_log(
        self, table_name, record_id, field_name, old_value, new_value, serial
    ):
        date = datetime.datetime.now().strftime("%d/%m/%Y")
        time = datetime.datetime.now().strftime("%H:%M:%S")
        self.cursor.execute(
            """
            INSERT INTO log_board (table_name, record_id, field_name, old_value, new_value, date, time, serial)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                table_name,
                record_id,
                field_name,
                old_value,
                new_value,
                date,
                time,
                serial,
            ),
        )
        self.conn.commit()

    def insert_data(self, row):
        self.cursor.execute(
            """
        INSERT INTO lavagem_placas (data, turno, hora, modelo, responsavel, linha_solicitante, blank_id, serial, fase)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            row,
        )
        self.conn.commit()

    def update_data(self, row, row_id):
        row_list = list(row)  # Convertendo a tupla em lista
        row_list.append(row_id)  # Adicionando o row_id à lista
        self.cursor.execute(
            """
    UPDATE lavagem_placas
    SET data = ?, turno = ?, hora = ?, modelo = ?, responsavel = ?, linha_solicitante = ?, blank_id = ?, serial = ?, fase = ?
    WHERE id = ?
    """,
            row_list,
        )
        self.conn.commit()

    def delete_data(self, row_id):
        self.cursor.execute(
            """
        DELETE FROM lavagem_placas WHERE id = ?
        """,
            (row_id,),
        )
        self.conn.commit()

    def fetch_all_data(self):
        self.cursor.execute(
            """
        SELECT * FROM lavagem_placas
        """
        )
        return self.cursor.fetchall()

    def close(self):
        self.conn.close()


class InserirSeriais(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Inserir Seriais")
        self.layout = QVBoxLayout()

        self.serials_inserted_per_entry = {}  # Inicializando a variável aqui
        self.blank_ids = []  # Inicializando a variável aqui

        self.finalized = False

        self.num_blanks_label = QLabel("Quantos blanks?")
        self.num_blanks_entry = QLineEdit("1")
        self.num_blanks_entry.textChanged.connect(self.update_serial_fields)
        self.num_seriais_label = QLabel("Quantos Seriais?")
        self.num_seriais_entry = QLineEdit()
        self.seriais_entries = []
        self.finalizar_button = QPushButton("FINALIZAR")
        self.finalizar_button.clicked.connect(self.finalizar)

        # Ao marcar a caixinha e finalizar, a quantidade inserida, irá virar SN, automaticamente.
        self.no_serial_checkbox = QCheckBox("Placas sem Serial")

        self.layout.addWidget(self.num_blanks_label)
        self.layout.addWidget(self.num_blanks_entry)
        self.layout.addWidget(self.num_seriais_label)
        self.layout.addWidget(self.num_seriais_entry)
        self.layout.addWidget(self.no_serial_checkbox)  # Adicionado de volta
        self.layout.addWidget(self.finalizar_button)

        # Estilizando os elementos
        self.setStyleSheet(
            """
            QLabel {
                font-size: 16px;
                color: #333;
            }
            QLineEdit, QPushButton {
                font-size: 14px;
                padding: 5px;
                background-color: #FFF;
                border: 1px solid #CCC;
                border-radius: 3px;
            }
            QCheckBox {
                font-size: 14px;
                color: #333;
            }
            QPushButton {
                padding: 5px 10px;
                background-color: #007a0d;
                color: #FFF;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """
        )

        self.setLayout(self.layout)

        self.serials_inserted = 0
        self.update_serial_fields()  # Chame isso para inicializar o campo para 1 blank

    def update_serial_fields(self):
        try:
            num_blanks = int(self.num_blanks_entry.text())
        except ValueError:
            return  # Ignora se o valor não for um número

        for entry in self.seriais_entries:
            self.layout.removeWidget(entry)
            entry.deleteLater()
        self.seriais_entries = []
        for i in range(num_blanks):
            entry = QLineEdit()
            entry.setPlaceholderText(f"Seriais para o blank {i+1}")
            self.layout.insertWidget(self.layout.count() - 1, entry)
            self.seriais_entries.append(entry)

    def keyPressEvent(self, event):
        for seriais_entry in self.seriais_entries:
            if (
                event.key() in (Qt.Key_Tab, Qt.Key_Enter, Qt.Key_Return)
                and seriais_entry.hasFocus()
            ):
                self.add_semicolon(seriais_entry)
                event.ignore()
                return
        super().keyPressEvent(event)

    def get_num_seriais(self):
        return int(self.num_seriais_entry.text())

    def add_semicolon(self, entry):
        if entry not in self.serials_inserted_per_entry:
            self.serials_inserted_per_entry[entry] = (
                0  # Inicializa se ainda não existir
            )

        max_seriais = int(self.num_seriais_entry.text())
        if self.serials_inserted_per_entry[entry] < max_seriais:
            current_text = entry.text()
            entry.setText(current_text + ";")
            self.serials_inserted_per_entry[entry] += 1
        else:
            QMessageBox.warning(
                self, "Limite Atingido", "Número máximo de seriais inseridos."
            )

    def finalizar(self):
        self.num_blanks = int(self.num_blanks_entry.text())
        self.seriais_per_blank = [
            entry.text().strip(";") for entry in self.seriais_entries
        ]

        # Verificação adicional para garantir que o serial comece com uma letra ou número
        if any(
            seriais and not seriais[0].isalnum() for seriais in self.seriais_per_blank
        ):
            QMessageBox.warning(
                self,
                "Erro de Entrada",
                "O serial deve começar com pelo menos um caractere ou número.",
            )
            return

        self.accept()
        for _ in range(self.num_blanks):
            blank_id = "".join(
                random.choices(string.ascii_letters + string.digits, k=6)
            )
            self.blank_ids.append(blank_id)

        if self.no_serial_checkbox.isChecked():  # Adicionado de volta
            self.seriais_per_blank = [
                "SN;" * int(self.num_seriais_entry.text())
            ] * self.num_blanks
        else:
            self.seriais_per_blank = [
                entry.text().strip(";") for entry in self.seriais_entries
            ]

        self.accept()


class ConsultResultWindow(QDialog):
    def __init__(self, results, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Resultados da Consulta")

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(
            [
                "DATA",
                "TURNO",
                "HORA",
                "MODELO",
                "RESPONSAVEL",
                "LINHA SOLICITANTE",
                "BLANK ID",
                "SERIAL",
                "FASE",
            ]
        )

        # Defina alguns estilos para a tabela
        self.table.setStyleSheet(
            "QTableWidget { background-color: #f0f0f0; border: 1px solid #ccc; }"
        )
        self.table.horizontalHeader().setStyleSheet(
            "QHeaderView::section { background-color: #333; color: white; border: 1px solid #ccc; }"
        )
        self.table.setFont(QFont("Arial", 10))
        self.table.horizontalHeader().setStretchLastSection(True)

        self.populate_table(results)

        self.layout.addWidget(self.table)

        # Ajustar tamanho da janela e da tabela
        self.resize(900, 700)
        self.table.resizeColumnsToContents()

    def populate_table(self, results):
        self.table.setRowCount(len(results))
        for i, row in enumerate(results):
            for j, item in enumerate(row):
                self.table.setItem(i, j, QTableWidgetItem(item))


class ConsultWindow(QDialog):
    def __init__(self, data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Consulta")

        self.db_manager = DatabaseManager()
        responsaveis = self.db_manager.fetch_all_responsaveis()
        modelos = self.db_manager.fetch_all_modelos()

        self.data = data
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.data_label = QLabel("DATA:")
        self.data_entry = QLineEdit()
        self.turno_label = QLabel("TURNO:")
        self.turno_combo = QComboBox()
        self.turno_combo.addItem("")
        self.turno_combo.addItems(["1°", "2°", "3°"])
        self.hora_label = QLabel("HORA:")
        self.hora_entry = QLineEdit()
        self.modelo_label = QLabel("MODELO:")
        self.modelo_combo = QComboBox()
        self.serial_label = QLabel("SERIAL:")
        self.serial_entry = QLineEdit()
        self.responsavel_label = QLabel("RESPONSAVEL:")
        self.responsavel_combo = QComboBox()
        self.linha_solicitante_label = QLabel("LINHA SOLICITANTE:")
        self.linha_solicitante_combo = QComboBox()
        self.linha_solicitante_combo.addItems(
            [
                " ",
                "Linha Manaus",
                "Linha Manicoré",
                "Linha Tefé",
                "Linha Coari",
                "Linha Tabatinga",
                "Linha Manacapuru",
                "Linha Autazes",
                "Linha Itacoatiara",
                "Linha Codajás",
                "Linha Parintins",
                "Linha Maués",
                "Linha Japurá",
            ]
        )
        self.fase_label = QLabel("FASE:")
        self.fase_combo = QComboBox()
        self.fase_combo.addItem("")
        self.fase_combo.addItems(["1°", "2°"])
        self.consult_button = QPushButton("Fazer Consulta")
        self.consult_button.clicked.connect(self.perform_consult)
        self.consult_button.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; padding: 10px 20px; border: none; border-radius: 5px; font-size: 16px; }"
            "QPushButton:hover { background-color: #45a049; }"
        )

        self.placa_sem_serial_checkbox = QCheckBox("Placa Sem Serial")
        self.placa_sem_serial_checkbox.stateChanged.connect(self.toggle_fields)

        self.layout.addWidget(self.data_label)
        self.layout.addWidget(self.data_entry)
        self.layout.addWidget(self.turno_label)
        self.layout.addWidget(self.turno_combo)
        self.layout.addWidget(self.hora_label)
        self.layout.addWidget(self.hora_entry)
        self.layout.addWidget(self.modelo_label)
        self.layout.addWidget(self.modelo_combo)
        self.layout.addWidget(self.serial_label)
        self.layout.addWidget(self.serial_entry)
        self.layout.addWidget(self.responsavel_label)
        self.layout.addWidget(self.responsavel_combo)
        self.layout.addWidget(self.linha_solicitante_label)
        self.layout.addWidget(self.linha_solicitante_combo)
        self.layout.addWidget(self.fase_label)
        self.layout.addWidget(self.fase_combo)
        self.layout.addWidget(self.placa_sem_serial_checkbox)
        self.layout.addWidget(self.consult_button)

        self.responsavel_combo.addItems([resp[1] for resp in responsaveis])
        self.modelo_combo.addItems([modelo[1] for modelo in modelos])

        self.filtered_data = []

        # Desabilita os campos por padrão
        self.enable_or_disable_fields(False)

    def toggle_fields(self, state):
        if state == Qt.Checked:
            self.enable_or_disable_fields(True)
        else:
            self.enable_or_disable_fields(False)

    def enable_or_disable_fields(self, enable):
        self.data_entry.setEnabled(enable)
        self.turno_combo.setEnabled(enable)
        self.hora_entry.setEnabled(enable)
        self.modelo_combo.setEnabled(enable)
        self.responsavel_combo.setEnabled(enable)
        self.linha_solicitante_combo.setEnabled(enable)
        self.fase_combo.setEnabled(enable)

    def perform_consult(self):
        query_data = [
            self.data_entry.text() if self.data_entry.isEnabled() else "",
            self.turno_combo.currentText() if self.turno_combo.isEnabled() else "",
            self.hora_entry.text() if self.hora_entry.isEnabled() else "",
            self.modelo_combo.currentText() if self.modelo_combo.isEnabled() else "",
            (
                self.responsavel_combo.currentText()
                if self.responsavel_combo.isEnabled()
                else ""
            ),
            (
                self.linha_solicitante_combo.currentText()
                if self.linha_solicitante_combo.isEnabled()
                else ""
            ),
            "",
            self.serial_entry.text(),  # Serial sempre habilitado
        ]

        self.db_manager = DatabaseManager()
        all_data = self.db_manager.fetch_all_data()

        # Se todos os campos estiverem vazios, puxa todos os registros
        if all(not query for query in query_data):
            self.filtered_data = [row[1:] for row in all_data]
        else:
            self.filtered_data = [
                row[1:]
                for row in all_data
                if all(
                    query == item or query == ""
                    for query, item in zip(query_data, row[1:])
                )
            ]

        result_window = ConsultResultWindow(self.filtered_data, self)
        result_window.exec_()


class EditWindow(QDialog):
    def __init__(self, row_data, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Editar Registro")
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Definindo um estilo CSS para a janela
        self.setStyleSheet(
            """
            QDialog {
                background-color: #F5F5F5;
            }
            QLabel {
                font-size: 14px;
                color: #333;
            }
            QLineEdit, QComboBox {
                font-size: 14px;
                padding: 5px;
                background-color: #FFF;
                border: 1px solid #CCC;
                border-radius: 3px;
            }
            QPushButton {
                font-size: 14px;
                padding: 5px 10px;
                background-color: #01700d;
                color: #FFF;
                border: none;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """
        )

        # Inicializando db_manager
        self.db_manager = DatabaseManager()

        # Copiando os dados da linha selecionada
        self.row_data = list(row_data)

        # Criando campos de entrada para edição
        labels = [
            "DATA",
            "TURNO",
            "HORA",
            "MODELO",
            "RESPONSAVEL",
            "LINHA SOLICITANTE",
            "BLANK ID",
            "SERIAL",
            "FASE",
        ]
        self.entries = []
        for i, label in enumerate(labels):
            label_widget = QLabel(label)

            if label == "MODELO":
                entry_widget = QComboBox()

                # Carregar modelos do banco de dados
                modelos = self.db_manager.fetch_all_modelos()
                entry_widget.addItems([modelo[1] for modelo in modelos])
                entry_widget.setCurrentText(self.row_data[i])

            if label == "LINHA SOLICITANTE":
                entry_widget = QComboBox()
                entry_widget.addItems(
                    [
                        "Linha Manaus",
                        "Linha Tefé",
                        "Linha Coari",
                        "Linha Tabatinga",
                        "Linha Autazes",
                        "Linha Itacoatiara",
                        "Linha Codajás",
                        "Linha Parintins",
                        "Linha Manacapuru",
                        "Linha Maués",
                        "Linha Japurá",
                    ]
                )
                entry_widget.setCurrentText(self.row_data[i])

            elif label == "TURNO":
                entry_widget = QComboBox()
                entry_widget.addItems(["1°", "2°", "3°"])
                entry_widget.setCurrentText(self.row_data[i])
            elif label == "FASE":
                entry_widget = QComboBox()
                entry_widget.addItems(["1°", "2°"])
                entry_widget.setCurrentText(self.row_data[i])
            else:
                entry_widget = QLineEdit(self.row_data[i])

            if label == "SERIAL":  # Bloquear a edição do SERIAL
                entry_widget.setReadOnly(True)
            if label == "RESPONSAVEL":  # Bloquear edição do RESPONSAVEL
                entry_widget.setReadOnly(True)
            if label == "MODELO":  # Bloquear edição do MODELO
                entry_widget.setReadOnly(True)
            if label == "DATA":  # Bloquear edição do MODELO
                entry_widget.setReadOnly(True)
            if label == "HORA":  # Bloquear edição do HORÁRIO
                entry_widget.setReadOnly(True)
            if label == "BLANK ID":  # Bloquear edição do ID
                entry_widget.setReadOnly(True)

            self.layout.addWidget(label_widget)
            self.layout.addWidget(entry_widget)
            self.entries.append(entry_widget)

        self.finalizar_button = QPushButton("FINALIZAR")
        self.finalizar_button.clicked.connect(self.finalizar)
        self.layout.addWidget(self.finalizar_button)

    def finalizar(self):
        # Atualizando os dados da linha com as edições
        for i, entry in enumerate(self.entries):
            if isinstance(entry, QComboBox):
                self.row_data[i] = entry.currentText()
            else:
                self.row_data[i] = entry.text()

        # Gerando um novo ID para o registro editado
        new_blank_id = "".join(
            random.choices(string.ascii_letters + string.digits, k=6)
        )
        self.row_data[6] = new_blank_id

        self.accept()


class DateRangeDialog(QDialog):
    def __init__(self, parent=None):
        super(DateRangeDialog, self).__init__(parent)
        self.setWindowTitle("Selecionar Intervalo de Datas")

        layout = QVBoxLayout()

        self.label = QLabel("Selecione o intervalo de datas:")
        layout.addWidget(self.label)

        # Estilizando o QLabel
        self.label.setStyleSheet(
            """
            font-size: 16px;
            color: #333;
        """
        )

        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDate(datetime.datetime.now().date())

        # Estilizando o QDateEdit
        self.start_date.setStyleSheet(
            """
            font-size: 14px;
            padding: 5px;
            background-color: #FFF;
            border: 1px solid #CCC;
            border-radius: 3px;
        """
        )

        layout.addWidget(self.start_date)

        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDate(datetime.datetime.now().date())

        # Estilizando o QDateEdit
        self.end_date.setStyleSheet(
            """
            font-size: 14px;
            padding: 5px;
            background-color: #FFF;
            border: 1px solid #CCC;
            border-radius: 3px;
        """
        )

        layout.addWidget(self.end_date)

        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)

        # Estilizando o QPushButton
        self.ok_button.setStyleSheet(
            """
            font-size: 16px;
            padding: 5px 10px;
            background-color: #007a0d;
            color: #FFF;
            border: none;
            border-radius: 3px;
        """
        )

        layout.addWidget(self.ok_button)

        self.setLayout(layout)

        # Definindo valores padrão para start_date e end_date
        self.start_date.setDate(datetime.datetime.now().date())
        self.end_date.setDate(datetime.datetime.now().date())


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CLP - Controle de Lavagem de Placas")

        # Removida a linha de FixedSize para permitir redimensionamento
        self.setFixedSize(1200, 1000)

        # Configura o widget para o background
        self.backgroundWidget = QLabel(self)
        dir_path = os.path.dirname(os.path.realpath(__file__))
        background_image_path = os.path.join(
            dir_path, "background.jpg"
        )  # Caminho relativo para background.png
        self.backgroundWidget.setPixmap(QPixmap(background_image_path))
        self.backgroundWidget.setScaledContents(True)
        self.backgroundWidget.setGeometry(0, 0, 1200, 1000)

        # Define a opacidade da imagem de fundo
        opacidade = QGraphicsOpacityEffect(self.backgroundWidget)
        opacidade.setOpacity(
            0.3
        )  # Ajuste este valor para a opacidade desejada (0.0 a 1.0)
        self.backgroundWidget.setGraphicsEffect(opacidade)

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Opção para iniciar em tela cheia ou permitir maximizar
        # self.showMaximized()

        self.db_manager = DatabaseManager()

        self.inserted_seriais = None  # Adicionado para armazenar os seriais inseridos
        self.num_blanks = None

        # Estilos
        label_style = "QLabel { font-weight: bold; font-size: 14px; }"
        entry_style = (
            "QLineEdit { border: 1px solid #A9A9A9; padding: 5px; font-size: 14px; }"
        )
        combo_style = (
            "QComboBox { border: 1px solid #A9A9A9; padding: 2px; font-size: 14px; }"
        )
        button_style = "QPushButton { background-color: #4CAF50; color: white; padding: 10px 2px; border: none; border-radius: 5px; font-size: 16px; } QPushButton:hover { background-color: #45a049; }"

        # Inserir rótulos e campos de entrada de texto
        self.data_label = QLabel("DATA:")
        self.data_label.setStyleSheet(label_style)
        self.data_entry = QLineEdit()
        self.data_entry.setStyleSheet(entry_style)

        self.turno_label = QLabel("TURNO:")
        self.turno_label.setStyleSheet(label_style)
        self.turno_combo = QComboBox()
        self.turno_combo.setStyleSheet(combo_style)
        self.turno_combo.addItems(["1°", "2°", "3°"])

        self.hora_label = QLabel("HORA:")
        self.hora_label.setStyleSheet(label_style)
        self.hora_entry = QLineEdit()
        self.hora_entry.setStyleSheet(entry_style)

        self.modelo_label = QLabel("MODELO:")
        self.modelo_label.setStyleSheet(label_style)
        self.modelo_combo = QComboBox()
        self.modelo_combo.setStyleSheet(combo_style)

        self.inserir_seriais_button = QPushButton("Inserir Seriais ao Blank")
        self.inserir_seriais_button.setStyleSheet(button_style)
        self.inserir_seriais_button.clicked.connect(self.abrir_inserir_seriais)

        self.responsavel_label = QLabel("OPERADOR:")
        self.responsavel_label.setStyleSheet(label_style)
        self.responsavel_combo = QComboBox()
        self.responsavel_combo.setStyleSheet(combo_style)

        self.linha_solicitante_label = QLabel("LINHA SOLICITANTE:")
        self.linha_solicitante_label.setStyleSheet(label_style)
        self.linha_solicitante_combo = QComboBox()
        self.linha_solicitante_combo.setStyleSheet(combo_style)
        self.linha_solicitante_combo.addItems(
            [
                "Linha Manaus",
                "Linha Manicoré",
                "Linha Tefé",
                "Linha Coari",
                "Linha Tabatinga",
                "Linha Manacapuru",
                "Linha Autazes",
                "Linha Itacoatiara",
                "Linha Codajás",
                "Linha Parintins",
                "Linha Maués",
                "Linha Japurá",
            ]
        )

        self.fase_label = QLabel("FASE:")
        self.fase_label.setStyleSheet(label_style)
        self.fase_combo = QComboBox()
        self.fase_combo.setStyleSheet(combo_style)
        self.fase_combo.addItems(["1°", "2°"])

        # Botão de Inserção
        self.insert_button = QPushButton("Inserir")
        self.insert_button.setIcon(
            QIcon(icon_path("correto.png"))
        )  # Ícone para o botão Inserir
        self.insert_button.setEnabled(False)  # Desativado inicialmente
        self.insert_button.clicked.connect(self.insert_data)

        # Botão de Edição
        self.edit_button = QPushButton("Editar Registro")
        self.edit_button.setIcon(
            QIcon(icon_path("config.png"))
        )  # Ícone para o botão Editar
        self.edit_button.clicked.connect(self.edit_data)

        # Botão de Consulta
        self.consult_button = QPushButton("Consultar")
        self.consult_button.setIcon(
            QIcon(icon_path("lupa.png"))
        )  # Ícone para o botão Consultar
        self.consult_button.clicked.connect(self.open_consult_window)

        # Botão de Geração de Relatório Excel
        self.report_button = QPushButton("Gerar Relatório Excel")
        self.report_button.setIcon(
            QIcon(icon_path("rel.png"))
        )  # Ícone para o botão Gerar Relatório
        self.report_button.clicked.connect(self.generate_report)

        # Botão de Relatório Diário
        self.daily_report_button = QPushButton("Relatório Diário")
        self.daily_report_button.setIcon(
            QIcon(icon_path("gerar.png"))
        )  # Ícone para o botão Relatório Diário
        self.daily_report_button.clicked.connect(self.generate_daily_report)

        # Botão de Envio de Relatório
        self.send_report_button = QPushButton("Enviar Relatório")
        self.send_report_button.setIcon(
            QIcon(icon_path("enviar_rel.png"))
        )  # Ícone para o botão Enviar Relatório
        self.send_report_button.clicked.connect(self.open_send_report_window)

        # Botão Sobre
        self.about_button = QPushButton("Sobre")
        self.about_button.setIcon(
            QIcon(icon_path("sobre.png"))
        )  # Ícone para o botão Sobre
        self.about_button.clicked.connect(self.open_about_dialog)

        # Carregar modelos do banco de dados
        modelos = self.db_manager.fetch_all_modelos()
        self.modelo_combo.addItems([modelo[1] for modelo in modelos])
        responsaveis = self.db_manager.fetch_all_responsaveis()
        self.responsavel_combo.addItems(
            [responsavel[1] for responsavel in responsaveis]
        )

        # Estilização dos botões
        button_style = """
        QPushButton { background-color: #4CAF50; color: white; padding: 10px 20px; border: 2px solid #4CAF50; border-radius: 5px; font-size: 16px;  min-width: 80px; }
        QPushButton:hover { background-color: #45a049; }
        QPushButton:disabled { background-color: #A5D6A7; }
        """
        self.insert_button.setStyleSheet(button_style)
        self.edit_button.setStyleSheet(
            button_style.replace("#4CAF50", "#FF0000").replace("#45a049", "#D32F2F")
        )
        self.consult_button.setStyleSheet(
            button_style.replace("#4CAF50", "#808080").replace("#45a049", "#696969")
        )
        self.report_button.setStyleSheet(
            button_style.replace("#4CAF50", "#D3D3D3").replace("#45a049", "#C0C0C0")
        )
        self.daily_report_button.setStyleSheet(
            button_style.replace("#4CAF50", "#696969").replace("#45a049", "#505050")
        )
        self.send_report_button.setStyleSheet(
            button_style.replace("#4CAF50", "#696969").replace("#45a049", "#505050")
        )
        self.about_button.setStyleSheet(
            button_style.replace("#4CAF50", "#00177f").replace("#45a049", "#505050")
        )

        # Adicionar ao layout
        self.layout.addWidget(self.data_label)
        self.layout.addWidget(self.data_entry)
        self.layout.addWidget(self.turno_label)
        self.layout.addWidget(self.turno_combo)
        self.layout.addWidget(self.hora_label)
        self.layout.addWidget(self.hora_entry)
        self.layout.addWidget(self.modelo_label)
        self.layout.addWidget(self.modelo_combo)

        self.layout.addWidget(self.responsavel_label)
        self.layout.addWidget(self.responsavel_combo)
        self.layout.addWidget(self.linha_solicitante_label)
        self.layout.addWidget(self.linha_solicitante_combo)
        self.layout.addWidget(self.fase_label)
        self.layout.addWidget(self.fase_combo)
        self.layout.addWidget(self.inserir_seriais_button)
        self.layout.addWidget(self.insert_button)
        self.layout.addWidget(self.edit_button)
        self.layout.addWidget(self.consult_button)
        self.layout.addWidget(self.about_button)

        # Configuração da tabela
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(
            [
                "DATA",
                "TURNO",
                "HORA",
                "MODELO",
                "OPERADOR",
                "LINHA SOLICITANTE",
                "BLANK ID",
                "SERIAL",
                "FASE",
            ]
        )
        self.layout.addWidget(self.table)

        # Ajustar largura das colunas
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        # Estilização da tabela
        self.table.setStyleSheet(
            "QTableWidget {"
            "background-color: #FFFFFF;"
            "border: 1px solid #A9A9A9;"
            "gridline-color: #A9A9A9;"
            "}"
            "QTableWidget::item:selected {"
            "background-color: #87CEEB;"  # Cor de fundo para itens selecionados
            "color: black;"  # Cor do texto para itens selecionados
            "}"
            "QHeaderView::section {"
            "background-color: #006400;"
            "color: white;"
            "padding: 10px;"
            "border: 1px solid #A9A9A9;"
            "font-size: 13px;"
            "font-weight: bold;"
            "}"
            "QTableWidget::horizontalHeader {"
            "background-color: #4CAF50;"
            "}"
            "QScrollBar:vertical {"
            "width: 18px;"
            "}"
        )

        # Centralizar a tabela no layout
        self.layout.setAlignment(self.table, Qt.AlignCenter)
        self.layout.addWidget(self.table)

        # Estilização da seleção
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)

        # Botão para gerar relatório
        self.layout.addWidget(self.report_button)

        self.layout.addWidget(self.daily_report_button)

        self.layout.addWidget(self.send_report_button)

        # Armazenar dados inseridos
        self.data = []

        # Preencher campos DATA e HORA automaticamente
        current_date = datetime.datetime.now().strftime("%d/%m/%Y")
        self.data_entry.setText(current_date)
        current_time = datetime.datetime.now().strftime("%H:%M:%S")
        self.hora_entry.setText(current_time)

        # Atualizar hora a cada segundo
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time)
        self.timer.start(1000)

        # Carregar dados salvos
        self.load_data()

    def open_about_dialog(self):
        # Criar uma janela de diálogo para mostrar as informações
        dialog = QDialog(self)
        dialog.setWindowTitle("Sobre")
        dialog.setStyleSheet(
            "background-color: #f2f2f2; font-family: Arial;"
        )  # Estilo da janela

        layout = QVBoxLayout()
        about_label = QLabel(
            "Sistema desenvolvido pela empresa HyperTech.\n"
            "\nGithub: https://github.com/HyperTechDevelopment\n"
            "\nCLP - CONTROLE DE LAVAGEM DE PLACAS"
        )
        about_label.setStyleSheet(
            "color: #333; padding: 20px; font-size: 14px;"
        )  # Estilo do rótulo
        layout.addWidget(about_label)

        dialog.setLayout(layout)
        dialog.exec_()

    def open_send_report_window(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Enviar Relatório")
        layout = QVBoxLayout()

        # Grupo de seleção de coordenadores
        coord_group = QGroupBox("Selecione os Coordenadores")
        coord_layout = QFormLayout()
        coord1 = QCheckBox("Coordenador (E-mail do coordenador)")
        coord2 = QCheckBox("Coordenador 2 (E-mail do coordenador)")
        coord3 = QCheckBox("Coordenador 3 (E-mail do coordenador)")
        coord4 = QCheckBox("Coordenador 4 (E-mail do coordenador)")
        coord5 = QCheckBox("Coordenador 5 (E-mail do coordenador)")
        coord6 = QCheckBox("Coordenador 6 (E-mail do coordenador)")
        coord7 = QCheckBox("Coordenador 7 (E-mail do coordenador)")
        coord8 = QCheckBox("Coordenador 8 (E-mail do coordenador)")
        coord_layout.addRow(coord1)
        coord_layout.addRow(coord2)
        coord_layout.addRow(coord3)
        coord_layout.addRow(coord4)
        coord_layout.addRow(coord5)
        coord_layout.addRow(coord6)
        coord_layout.addRow(coord7)
        coord_layout.addRow(coord8)
        coord_group.setLayout(coord_layout)
        layout.addWidget(coord_group)

        # Botão para selecionar o arquivo
        select_file_button = QPushButton("Selecionar Arquivo")
        select_file_button.clicked.connect(lambda: self.select_file(dialog))
        layout.addWidget(select_file_button)

        # Botão para enviar o relatório
        send_button = QPushButton("Enviar")
        send_button.clicked.connect(
            lambda: self.send_report(
                [coord1, coord2, coord3, coord4, coord5, coord6, coord7, coord8], dialog
            )
        )
        layout.addWidget(send_button)

        dialog.setLayout(layout)
        dialog.exec_()

    def select_file(self, dialog):
        options = QFileDialog.Options()
        dialog.file_name, _ = QFileDialog.getOpenFileName(
            dialog,
            "Selecionar Arquivo",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
            options=options,
        )

    def send_report(self, coords, dialog):
        selected_coords = [
            coord.text().split(" ")[1] for coord in coords if coord.isChecked()
        ]
        if not hasattr(dialog, "file_name"):
            QMessageBox.warning(
                dialog, "Arquivo não selecionado", "Por favor, selecione um arquivo."
            )
            return

        if not selected_coords:
            QMessageBox.warning(
                dialog,
                "Nenhum Coordenador Selecionado",
                "Por favor, selecione pelo menos um coordenador.",
            )
            return

        # Configurar o Outlook
        outlook = win32.Dispatch("outlook.application")
        mail = outlook.CreateItem(0)
        mail.To = "; ".join(selected_coords)
        mail.Subject = f'NOME DA SUA EMRPESA: Relatório Diário de Lavagem de Placas ({time.strftime("%d/%m/%y")})'

        # Determinar saudação com base na hora
        current_hour = int(time.strftime("%H"))
        if 7 <= current_hour <= 11:
            greeting = "Bom dia Srs,"
        elif 12 <= current_hour <= 17:
            greeting = "Boa tarde Srs,"
        else:
            greeting = "Boa noite Srs,"

        mail.Body = f"Olá\n\n{greeting}\n\nSegue em anexo o relatório diário de lavagem de placas,\n\nAtenciosamente,\n\nPRODUÇÃO-SMT"
        mail.Attachments.Add(dialog.file_name)
        mail.Send()

        QMessageBox.information(
            dialog, "Relatório Enviado", "O relatório foi enviado com sucesso."
        )
        dialog.close()

    def generate_daily_report(self):

        selected_date = simpledialog.askstring(
            "Selecionar Data",
            "Insira a data para o relatório (dd/mm/yyyy):",
            initialvalue=datetime.datetime.now().strftime("%d/%m/%Y"),
        )
        if not selected_date:
            return

        # Estrutura para armazenar os dados
        report_data = defaultdict(lambda: defaultdict(int))

        # Coletar dados do banco de dados para a data selecionada
        self.db_manager.cursor.execute(
            "SELECT * FROM lavagem_placas WHERE data = ?", (selected_date,)
        )
        all_data = self.db_manager.cursor.fetchall()

        if not all_data:
            messagebox.showwarning(
                "Nenhum Registro", "NÃO HÁ REGISTROS para a data selecionada."
            )
            return

        # Agrupar dados
        for row in all_data:
            data, turno, hora, modelo, responsavel, linha, blank_id, serial, fase = row[
                1:
            ]
            key = (modelo, linha, turno, fase, responsavel)
            report_data[key]["count"] += 1

        # Criar planilha
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Estilizar e mesclar A1:F1
        sheet.merge_cells("A1:F1")
        cell = sheet["A1"]
        cell.value = "RELATÓRIO DIÁRIO"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Arial Black", size=20, color="000000")
        cell.fill = PatternFill(
            start_color="9BBB59", end_color="9BBB59", fill_type="solid"
        )
        sheet.row_dimensions[1].height = 30

        # Adicionar e estilizar cabeçalhos
        headers = ["Modelo", "Linha", "Turno", "Fase", "Quantidade", "Responsável"]
        colors = ["FCE4D6", "FCE4D6", "FCE4D6", "FCE4D6", "FCE4D6", "FCE4D6"]
        for col_num, (header, color) in enumerate(zip(headers, colors), 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", size=12, bold=True)
            cell.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )  # Adicionando bordas
            sheet.column_dimensions[chr(64 + col_num)].width = 15

        # Preencher e estilizar dados
        row_num = 3
        for (modelo, linha, turno, fase, responsavel), data in report_data.items():
            for col_num, value in enumerate(
                [modelo, linha, turno, fase, data["count"], responsavel], 1
            ):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=11)
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
            row_num += 1

        # Salvar planilha
        current_date = datetime.datetime.now().strftime("%d-%m-%Y")
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Relatório Diário ({current_date})",
        )
        if filepath:
            wb.save(filepath)
            messagebox.showinfo(
                "Relatório Diário Gerado",
                f"O relatório diário foi gerado com sucesso e salvo em:\n\n{filepath}",
            )

    def abrir_inserir_seriais(self):
        inserir_seriais_dialog = InserirSeriais(self)
        result = inserir_seriais_dialog.exec_()
        if result == QDialog.Accepted:
            self.insert_button.setEnabled(True)
            self.inserted_seriais = inserir_seriais_dialog.seriais_per_blank
            self.blank_ids = inserir_seriais_dialog.blank_ids
            self.seriais_per_blank = inserir_seriais_dialog.seriais_per_blank

            current_date = self.data_entry.text()
            current_time = self.hora_entry.text()

            # Pegar os valores inseridos pelo usuário
            num_blanks = inserir_seriais_dialog.num_blanks
            num_seriais = inserir_seriais_dialog.get_num_seriais()
            seriais = inserir_seriais_dialog.seriais_per_blank

            # Gerar um ID aleatório para o blank
            blank_id = "".join(
                random.choices(string.ascii_letters + string.digits, k=6)
            )

    def update_time(self):
        # Atualizar a hora no campo de entrada de texto
        current_time = datetime.datetime.now().strftime("%H:%M:%S")
        self.hora_entry.setText(current_time)

    def load_data(self):
        self.data = self.db_manager.fetch_all_data()
        self.table.setRowCount(len(self.data))
        for i, row in enumerate(self.data):
            for j, item in enumerate(row[1:]):  # Excluindo o ID da linha
                self.table.setItem(i, j, QTableWidgetItem(str(item)))

    def insert_data(self):
        if (
            any(field.text() == "" for field in [self.data_entry, self.hora_entry])
            or self.responsavel_combo.currentText() == ""
            or not self.inserted_seriais
        ):
            QMessageBox.warning(
                self, "Campos Vazios", "Todos os campos devem ser preenchidos."
            )
            return

        current_date = self.data_entry.text()
        current_time = self.hora_entry.text()
        current_modelos = [
            self.modelo_combo.itemText(i) for i in range(self.modelo_combo.count())
        ]
        current_responsaveis = [
            self.responsavel_combo.itemText(i)
            for i in range(self.responsavel_combo.count())
        ]

        database_manager = DatabaseManager()

        for blank_id, seriais in zip(self.blank_ids, self.seriais_per_blank):
            seriais_for_this_blank = seriais.strip().split(";")
            seriais_for_this_blank = [
                serial for serial in seriais_for_this_blank if serial
            ]
            for serial in seriais_for_this_blank:
                row = [
                    current_date,
                    self.turno_combo.currentText(),
                    current_time,
                    self.modelo_combo.currentText(),
                    self.responsavel_combo.currentText(),
                    self.linha_solicitante_combo.currentText(),
                    blank_id,
                    serial,
                    self.fase_combo.currentText(),
                ]
                database_manager.insert_data(row)

        self.load_data()

        self.data_entry.clear()
        self.turno_combo.setCurrentIndex(0)
        self.hora_entry.clear()
        self.modelo_combo.setCurrentIndex(0)
        self.responsavel_combo.setCurrentIndex(0)
        self.linha_solicitante_combo.setCurrentIndex(0)

        self.inserted_seriais = None
        self.insert_button.setEnabled(False)

        self.data_entry.setText(current_date)
        self.hora_entry.setText(current_time)

        self.modelo_combo.clear()
        self.modelo_combo.addItems(current_modelos)

        self.responsavel_combo.clear()
        self.responsavel_combo.addItems(current_responsaveis)

        QMessageBox.information(self, "Dados Inseridos", "Dados inseridos com sucesso!")

    def edit_data(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            old_row = list(self.data[selected_row])
            row_id = old_row[0]
            serial = old_row[8]  # O índice 8 corresponde à coluna 'serial'

            # Abrindo a janela de edição
            edit_window = EditWindow(old_row[1:], self)
            result = edit_window.exec_()
            if result == QDialog.Accepted:
                new_row = edit_window.row_data
                self.db_manager.update_data(new_row, row_id)

                # Inserir um log para cada campo alterado
                for i, (old_value, new_value) in enumerate(zip(old_row[1:], new_row)):
                    if old_value != new_value:
                        field_name = self.table.horizontalHeaderItem(i).text()
                        self.db_manager.insert_log(
                            "lavagem_placas",
                            row_id,
                            field_name,
                            old_value,
                            new_value,
                            serial,
                        )

                # Atualizando a tabela
                self.load_data()

                QMessageBox.information(
                    self, "Edição Concluída", "Registro editado com sucesso!"
                )

    def open_consult_window(self):
        # Abre a janela de consulta
        consult_window = ConsultWindow(self.data, self)
        consult_window.exec_()

    def generate_report(self):
        date_dialog = DateRangeDialog(self)
        result = date_dialog.exec_()

        if result == QDialog.Accepted:
            start_date = date_dialog.start_date.date().toPyDate()
            end_date = date_dialog.end_date.date().toPyDate()

            # Inicializando 'filtered_data' antes do bloco try-except
            filtered_data = []

            try:
                for row in self.data:
                    row_date = datetime.datetime.strptime(
                        str(row[1]).split()[0], "%d/%m/%Y"
                    ).date()
                    if start_date <= row_date <= end_date:
                        filtered_data.append(row)
            except ValueError as e:
                print(f"Erro ao converter data: {e}")

            # Processamento de 'filtered_data'
            if filtered_data:
                wb = openpyxl.Workbook()
                sheet = wb.active

                current_row = 4
                last_blank_id = None
                start_merge_row = current_row

                for i, row in enumerate(filtered_data):
                    current_blank_id = row[6]

                    if last_blank_id is not None and current_blank_id != last_blank_id:
                        for col in [1, 2, 3, 4, 5, 6]:
                            sheet.merge_cells(
                                start_row=start_merge_row,
                                end_row=current_row - 1,
                                start_column=col,
                                end_column=col,
                            )
                        start_merge_row = current_row

                    for j, item in enumerate(row[1:]):  # Começando do segundo elemento
                        sheet.cell(
                            row=current_row, column=j + 1, value=item
                        )  # Começando da coluna 2

                    last_blank_id = current_blank_id
                    current_row += 1

                # Mesclar células para o último grupo de 'BLANK ID', se necessário
                if last_blank_id is not None:
                    for col in [1, 2, 3, 4]:
                        sheet.merge_cells(
                            start_row=start_merge_row,
                            end_row=current_row - 1,
                            start_column=col,
                            end_column=col,
                        )

                # Definir o estilo de borda
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # Aplicar bordas a todas as células com dados
                for row in sheet.iter_rows(
                    min_row=3,
                    max_row=sheet.max_row,
                    min_col=1,
                    max_col=sheet.max_column,
                ):
                    for cell in row:
                        cell.border = thin_border

                # Ajustar o tamanho das colunas
                column_widths = {
                    "A": 15,  # DATA
                    "B": 10,  # TURNO
                    "C": 10,  # HORA
                    "D": 15,  # MODELO
                    "E": 20,  # RESPONSAVEL
                    "F": 28,  # LINHA SOLICITANTE
                    "G": 15,  # BLANK ID
                    "H": 15,  # SERIAL
                    "I": 10,  # FASE
                }

                for col, width in column_widths.items():
                    sheet.column_dimensions[col].width = width

                # Estilizar a planilha
                sheet.merge_cells("A1:I2")
                sheet["A1"].value = "CONTROLE DE LAVAGEM DE PLACAS"
                sheet["A1"].font = Font(name="Arial Black", size=22, color="00FFFFFF")
                sheet["A1"].fill = PatternFill(
                    start_color="063F67", end_color="063F67", fill_type="solid"
                )
                sheet["A1"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )

                # Adicionar rótulos para cada coluna
                headers = [
                    "DATA",
                    "TURNO",
                    "HORA",
                    "MODELO",
                    "RESPONSAVEL",
                    "LINHA SOLICITANTE",
                    "BLANK ID",
                    "SERIAL",
                    "FASE",
                ]
                for col_num, header in enumerate(headers, 1):
                    cell = sheet.cell(row=3, column=col_num)
                    cell.value = header
                    cell.font = Font(name="Calibri", size=11, bold=True)
                    cell.fill = PatternFill(
                        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Selecionar pasta de destino para o arquivo
                current_date = datetime.datetime.now().strftime("%d-%m-%Y")
                filepath, _ = QFileDialog.getSaveFileName(
                    self,
                    "Salvar Relatório",
                    f"Relatório-{current_date}",
                    "Excel Files (*.xlsx)",
                )
                if filepath:
                    # Salvar o arquivo
                    wb.save(filepath)
                    QMessageBox.information(
                        self,
                        "Relatório Gerado",
                        "O relatório foi gerado com sucesso e salvo em:\n\n" + filepath,
                    )

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self,
            "Fechar Sistema",
            "Tem certeza que deseja fechar o sistema?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            self.db_manager.close()
            event.accept()
        else:
            event.ignore()


def main():
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()


# DESENVOLVIDO POR:
# EMPRESA HyperTech ©
# SISTEMA DESENVOLVIDO PARA EMPRESA CALLIDUS, A FIM DE OTIMIZAR O PROCESSO DE CONTROLE DE LAVAGEM DE PLACAS.
